"""QuizBot – Telegram quiz bot powered by python-telegram-bot 20.7"""
from __future__ import annotations

import asyncio
import logging
import os
from pathlib import Path

from dotenv import load_dotenv
from telegram import InlineKeyboardButton, InlineKeyboardMarkup, Update
from telegram.ext import (
    Application,
    CallbackQueryHandler,
    CommandHandler,
    ContextTypes,
    MessageHandler,
    filters,
)

import tracker
from parser import parse_file
from quiz import QuizSession

load_dotenv()

logging.basicConfig(
    format="%(asctime)s  %(levelname)-8s  %(name)s — %(message)s",
    level=logging.INFO,
)
log = logging.getLogger(__name__)

BOT_TOKEN   = os.getenv("BOT_TOKEN", "")
SAMPLE_PATH = Path(__file__).parent / "sample.xlsx"

# ── Per-user in-memory state ───────────────────────────────────────────────────
_questions: dict[int, list[dict]] = {}   # user_id → loaded questions
_sessions:  dict[int, QuizSession] = {}  # user_id → active session
_shuffle:   dict[int, bool] = {}         # user_id → shuffle toggle
_uploading: set[int] = set()             # users currently expected to send a file


# ── Quiz engine helpers ────────────────────────────────────────────────────────

def _live_session(user_id: int) -> QuizSession | None:
    s = _sessions.get(user_id)
    return s if s and s.active else None


async def _send_question(ctx: ContextTypes.DEFAULT_TYPE, s: QuizSession) -> None:
    q = s.current_question
    if q is None:
        await _send_results(ctx, s)
        return

    s.question_id += 1
    qid = s.question_id
    s.awaiting_answer = True

    tl = q.get("time_limit", 15)
    kb = InlineKeyboardMarkup([
        [
            InlineKeyboardButton(f"A  {q['option_a']}", callback_data=f"ans_{qid}_a"),
            InlineKeyboardButton(f"B  {q['option_b']}", callback_data=f"ans_{qid}_b"),
        ],
        [
            InlineKeyboardButton(f"C  {q['option_c']}", callback_data=f"ans_{qid}_c"),
            InlineKeyboardButton(f"D  {q['option_d']}", callback_data=f"ans_{qid}_d"),
        ],
    ])

    text = (
        f"❓ *Question {s.current_index + 1}/{s.total}*\n"
        f"📂 _{q['category']}_   ⏱ {tl} s\n\n"
        f"{q['question']}"
    )
    msg = await ctx.bot.send_message(
        chat_id=s.chat_id,
        text=text,
        reply_markup=kb,
        parse_mode="Markdown",
    )
    s.current_msg_id = msg.message_id

    if s.timer_task and not s.timer_task.done():
        s.timer_task.cancel()
    s.timer_task = asyncio.create_task(_timeout(ctx, s, qid, tl))


async def _timeout(
    ctx: ContextTypes.DEFAULT_TYPE, s: QuizSession, qid: int, delay: int
) -> None:
    try:
        await asyncio.sleep(delay)
    except asyncio.CancelledError:
        return

    if not s.active or s.question_id != qid or not s.awaiting_answer:
        return

    s.awaiting_answer = False
    q = s.current_question
    if q is None:
        return

    correct      = q["correct_answer"]
    correct_text = q[f"option_{correct}"]

    s.record(q["category"], False)
    tracker.record_answer(s.user_id, q["category"], False)

    try:
        await ctx.bot.edit_message_reply_markup(
            chat_id=s.chat_id, message_id=s.current_msg_id, reply_markup=None
        )
    except Exception:
        pass

    if not s.active:
        return

    try:
        await ctx.bot.send_message(
            chat_id=s.chat_id,
            text=(
                f"⏰ *Time's up!*\n"
                f"✗ Wrong — correct answer was: *{correct.upper()}. {correct_text}*"
            ),
            parse_mode="Markdown",
        )
    except asyncio.CancelledError:
        return
    except Exception:
        pass

    if not s.active:
        return

    s.advance()
    if s.current_index >= s.total:
        await _send_results(ctx, s)
    elif s.active:
        await _send_question(ctx, s)


async def _send_results(ctx: ContextTypes.DEFAULT_TYPE, s: QuizSession) -> None:
    s.stop()
    _sessions.pop(s.user_id, None)

    pct_overall = s.score / s.total * 100 if s.total else 0
    lines: list[str] = [
        "🏁 *Quiz Complete!*\n",
        f"📊 *Score: {s.score}/{s.total}* ({pct_overall:.0f}%)\n",
        "\n📂 *Category breakdown:*",
    ]

    weak: list[str] = []
    for cat, data in sorted(s.cat_scores.items()):
        t, c = data["total"], data["correct"]
        pct  = c / t * 100 if t else 0
        icon = "✓" if pct >= 70 else "⚠️"
        if pct < 70:
            weak.append(cat)
        lines.append(f"  {icon} {cat}: {c}/{t} ({pct:.0f}%)")

    if weak:
        lines.append(f"\n⚠️ *Weak areas (<70%):* {', '.join(weak)}")
        lines.append("\nUse /weak to drill weak areas or /quiz to restart.")
    else:
        lines.append("\n🎉 No weak areas! Use /quiz to play again.")

    await ctx.bot.send_message(
        chat_id=s.chat_id,
        text="\n".join(lines),
        parse_mode="Markdown",
    )


# ── Command handlers ───────────────────────────────────────────────────────────

async def cmd_start(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> None:
    await update.message.reply_text(
        "👋 *Welcome to QuizBot!*\n\n"
        "*Quick start:*\n"
        "1. `/sample` — download the question template\n"
        "2. Fill it with your questions and send it here\n"
        "3. `/quiz` — start the full quiz\n\n"
        "*All commands:*\n"
        "`/quiz` — full quiz\n"
        "`/category [name]` — quiz one category\n"
        "`/weak` — drill categories below 70%\n"
        "`/shuffle` — toggle random question order\n"
        "`/score` — current score mid-quiz\n"
        "`/stop` — end quiz early\n"
        "`/sample` — download the template\n"
        "`/upload` — upload your quiz file",
        parse_mode="Markdown",
    )


async def cmd_sample(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> None:
    if not SAMPLE_PATH.exists():
        _create_sample_xlsx()
    with open(SAMPLE_PATH, "rb") as fh:
        await update.message.reply_document(
            document=fh,
            filename="sample.xlsx",
            caption=(
                "📋 Sample quiz file — 10 questions across 3 categories.\n"
                "Fill it with your content, then send the file here to upload it."
            ),
        )


async def cmd_upload(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> None:
    _uploading.add(update.effective_user.id)
    await update.message.reply_text(
        "📎 *Send your .xlsx or .csv file now.*\n\n"
        "Required columns: `question`, `option_a`, `option_b`, `option_c`, `option_d`, "
        "`correct_answer` (a/b/c/d), `category`\n"
        "Optional: `time_limit` (integer seconds, default 15)",
        parse_mode="Markdown",
    )


async def handle_document(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> None:
    uid  = update.effective_user.id
    doc  = update.message.document
    fname = doc.file_name or "quiz.xlsx"
    ext   = Path(fname).suffix.lower()

    if ext not in (".xlsx", ".xls", ".csv"):
        if uid in _uploading:
            await update.message.reply_text(
                "❌ Please send a *.xlsx* or *.csv* file.", parse_mode="Markdown"
            )
        return

    _uploading.discard(uid)
    status_msg = await update.message.reply_text("⏳ Parsing your file…")

    try:
        tg_file = await doc.get_file()
        raw     = bytes(await tg_file.download_as_bytearray())
        questions = parse_file(raw, fname)

        if not questions:
            await status_msg.edit_text(
                "❌ No valid questions found. Check the format and try again."
            )
            return

        _questions[uid] = questions
        cats = sorted({q["category"] for q in questions})
        cat_list = ", ".join(f"`{c}`" for c in cats)

        await status_msg.edit_text(
            f"✅ Loaded *{len(questions)} question{'s' if len(questions) != 1 else ''}* "
            f"across {len(cats)} categor{'ies' if len(cats) != 1 else 'y'}:\n"
            f"{cat_list}\n\n"
            f"Use /quiz to start!",
            parse_mode="Markdown",
        )
    except ValueError as exc:
        await status_msg.edit_text(f"❌ *Format error:*\n{exc}", parse_mode="Markdown")
    except Exception as exc:
        log.exception("File parse error")
        await status_msg.edit_text(f"❌ Unexpected error: {exc}")


async def cmd_quiz(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> None:
    uid = update.effective_user.id
    qs  = _questions.get(uid)

    if not qs:
        await update.message.reply_text(
            "❌ No questions loaded.\n"
            "Use /sample to get the template, fill it in, then send it here."
        )
        return
    if _live_session(uid):
        await update.message.reply_text("⚠️ A quiz is running. Use /stop to end it first.")
        return

    shuffle = _shuffle.get(uid, False)
    s = QuizSession(uid, update.effective_chat.id, qs, shuffle=shuffle)
    _sessions[uid] = s

    await update.message.reply_text(
        f"🚀 Starting quiz — *{len(qs)} questions* "
        f"({'🔀 shuffled' if shuffle else '📋 in order'})",
        parse_mode="Markdown",
    )
    await _send_question(ctx, s)


async def cmd_category(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> None:
    uid = update.effective_user.id
    qs  = _questions.get(uid)

    if not qs:
        await update.message.reply_text("❌ No questions loaded. Use /upload first.")
        return

    cats = sorted({q["category"] for q in qs})

    if not ctx.args:
        await update.message.reply_text(
            "*Available categories:*\n"
            + "\n".join(f"  • {c}" for c in cats)
            + "\n\nUsage: `/category <name>`",
            parse_mode="Markdown",
        )
        return

    name     = " ".join(ctx.args)
    filtered = [q for q in qs if q["category"].lower() == name.lower()]
    if not filtered:
        filtered = [q for q in qs if name.lower() in q["category"].lower()]
    if not filtered:
        await update.message.reply_text(
            f"❌ Category `{name}` not found.\n"
            f"Available: {', '.join(f'`{c}`' for c in cats)}",
            parse_mode="Markdown",
        )
        return

    if _live_session(uid):
        await update.message.reply_text("⚠️ A quiz is running. Use /stop first.")
        return

    shuffle    = _shuffle.get(uid, False)
    s          = QuizSession(uid, update.effective_chat.id, filtered, shuffle=shuffle)
    _sessions[uid] = s

    await update.message.reply_text(
        f"🚀 *{filtered[0]['category']}* quiz — *{len(filtered)} questions*",
        parse_mode="Markdown",
    )
    await _send_question(ctx, s)


async def cmd_shuffle(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> None:
    uid            = update.effective_user.id
    _shuffle[uid]  = not _shuffle.get(uid, False)
    state          = "ON 🔀" if _shuffle[uid] else "OFF 📋"
    await update.message.reply_text(f"Shuffle is now *{state}*", parse_mode="Markdown")


async def cmd_weak(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> None:
    uid = update.effective_user.id
    qs  = _questions.get(uid)

    if not qs:
        await update.message.reply_text("❌ No questions loaded. Use /upload first.")
        return

    weak_cats = tracker.get_weak_categories(uid)
    if not weak_cats:
        msg = (
            "📊 No quiz history yet — complete a quiz first."
            if not tracker.get_stats(uid)
            else "🎉 You're scoring ≥70% in every category!"
        )
        await update.message.reply_text(msg)
        return

    filtered = [q for q in qs if q["category"] in weak_cats]
    if not filtered:
        await update.message.reply_text(
            f"⚠️ Weak categories ({', '.join(weak_cats)}) "
            "have no matching questions in the loaded file."
        )
        return

    if _live_session(uid):
        await update.message.reply_text("⚠️ A quiz is running. Use /stop first.")
        return

    shuffle        = _shuffle.get(uid, False)
    s              = QuizSession(uid, update.effective_chat.id, filtered, shuffle=shuffle)
    _sessions[uid] = s

    await update.message.reply_text(
        f"💪 *Weak-area drill:* {', '.join(weak_cats)}\n*{len(filtered)} questions*",
        parse_mode="Markdown",
    )
    await _send_question(ctx, s)


async def cmd_score(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> None:
    uid = update.effective_user.id
    s   = _live_session(uid)

    if not s:
        await update.message.reply_text("❌ No quiz is running. Use /quiz to start one.")
        return

    answered = s.current_index
    lines    = [f"📊 *Score so far: {s.score}/{answered}*\n"]
    for cat, data in sorted(s.cat_scores.items()):
        t, c = data["total"], data["correct"]
        pct  = c / t * 100 if t else 0
        lines.append(f"  • {cat}: {c}/{t} ({pct:.0f}%)")
    lines.append(f"\n⏭ Next: question {answered + 1}/{s.total}")

    await update.message.reply_text("\n".join(lines), parse_mode="Markdown")


async def cmd_stop(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> None:
    uid = update.effective_user.id
    s   = _sessions.pop(uid, None)

    if not s or not s.active:
        await update.message.reply_text("❌ No quiz is running.")
        return

    answered = s.current_index
    s.stop()
    await update.message.reply_text(
        f"🛑 Quiz stopped.\nScore: *{s.score}/{answered}* answered.",
        parse_mode="Markdown",
    )


async def answer_callback(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    await query.answer()

    uid = query.from_user.id
    try:
        _, qid_s, letter = query.data.split("_", 2)
        qid = int(qid_s)
    except Exception:
        return

    s = _sessions.get(uid)
    if not s or not s.active or s.question_id != qid or not s.awaiting_answer:
        return  # stale tap or no session

    if s.timer_task and not s.timer_task.done():
        s.timer_task.cancel()
    s.awaiting_answer = False

    q            = s.current_question
    correct      = q["correct_answer"]
    correct_text = q[f"option_{correct}"]
    is_correct   = letter == correct

    s.record(q["category"], is_correct)
    tracker.record_answer(uid, q["category"], is_correct)

    try:
        await query.edit_message_reply_markup(reply_markup=None)
    except Exception:
        pass

    feedback = (
        "✅ *Correct!*"
        if is_correct
        else f"❌ *Wrong* — correct answer was: *{correct.upper()}. {correct_text}*"
    )
    await query.message.reply_text(feedback, parse_mode="Markdown")

    s.advance()
    if s.current_index >= s.total:
        await _send_results(ctx, s)
    elif s.active:
        await _send_question(ctx, s)


# ── Sample XLSX generator ──────────────────────────────────────────────────────

def _create_sample_xlsx() -> None:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Questions"

    headers = [
        "question", "option_a", "option_b", "option_c", "option_d",
        "correct_answer", "category", "time_limit",
    ]
    ws.append(headers)

    rows = [
        # General Knowledge (4 questions)
        (
            "What is the capital of France?",
            "London", "Berlin", "Paris", "Madrid",
            "c", "General Knowledge", 15,
        ),
        (
            "How many continents are there on Earth?",
            "5", "6", "7", "8",
            "c", "General Knowledge", 15,
        ),
        (
            "Which planet is known as the Red Planet?",
            "Venus", "Jupiter", "Saturn", "Mars",
            "d", "General Knowledge", 10,
        ),
        (
            "How many sides does a hexagon have?",
            "5", "6", "7", "8",
            "b", "General Knowledge", 10,
        ),
        # Science (3 questions)
        (
            "What is the chemical symbol for water?",
            "CO2", "O2", "H2O", "NaCl",
            "c", "Science", 12,
        ),
        (
            "Approximate speed of light in km/s?",
            "150,000", "300,000", "450,000", "600,000",
            "b", "Science", 15,
        ),
        (
            "Which gas do plants absorb during photosynthesis?",
            "Oxygen", "Nitrogen", "Carbon Dioxide", "Hydrogen",
            "c", "Science", 12,
        ),
        # History (3 questions)
        (
            "In what year did World War II end?",
            "1943", "1944", "1945", "1946",
            "c", "History", 15,
        ),
        (
            "Who was the first President of the United States?",
            "Abraham Lincoln", "Thomas Jefferson", "George Washington", "John Adams",
            "c", "History", 15,
        ),
        (
            "In which city was the ancient Lighthouse, one of the Seven Wonders?",
            "Rome", "Athens", "Alexandria", "Carthage",
            "c", "History", 12,
        ),
    ]
    for r in rows:
        ws.append(list(r))

    # Style header row
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
    for cell in ws[1]:
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = Alignment(horizontal="center")

    col_widths = [52, 22, 22, 22, 22, 16, 20, 12]
    for col, width in zip(ws.columns, col_widths):
        ws.column_dimensions[col[0].column_letter].width = width

    wb.save(SAMPLE_PATH)
    log.info("Created %s", SAMPLE_PATH)


# ── Entry point ────────────────────────────────────────────────────────────────

def main() -> None:
    if not BOT_TOKEN or BOT_TOKEN == "your_bot_token_here":
        raise SystemExit(
            "BOT_TOKEN is not set.\n"
            "Get a token from @BotFather on Telegram and add it to .env"
        )

    tracker.init_db()
    if not SAMPLE_PATH.exists():
        _create_sample_xlsx()

    app = Application.builder().token(BOT_TOKEN).build()

    app.add_handler(CommandHandler("start",    cmd_start))
    app.add_handler(CommandHandler("sample",   cmd_sample))
    app.add_handler(CommandHandler("upload",   cmd_upload))
    app.add_handler(CommandHandler("quiz",     cmd_quiz))
    app.add_handler(CommandHandler("category", cmd_category))
    app.add_handler(CommandHandler("shuffle",  cmd_shuffle))
    app.add_handler(CommandHandler("weak",     cmd_weak))
    app.add_handler(CommandHandler("score",    cmd_score))
    app.add_handler(CommandHandler("stop",     cmd_stop))
    app.add_handler(CallbackQueryHandler(answer_callback, pattern=r"^ans_"))
    app.add_handler(MessageHandler(filters.Document.ALL, handle_document))

    log.info("QuizBot is polling…")
    app.run_polling(allowed_updates=Update.ALL_TYPES)


if __name__ == "__main__":
    main()
